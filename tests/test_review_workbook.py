import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


class ReviewWorkbookTests(unittest.TestCase):
    def test_builds_portable_workbook_with_formulas_and_text_identifiers(self):
        review = {
            "metadata": {"reportCodes": ["ER-DEMO"], "reportTotal": 25.5, "failedExpenses": 0},
            "categories": [{"expenseType": "过路费", "applicationAmount": 30}],
            "rows": [{
                "fileName": "invoice.ofd", "format": "OFD", "documentType": "发票",
                "invoiceNumber": "001234567890", "suggestedCategory": "过路费",
                "confirmedCategory": "过路费", "reportGroup": "差旅报销",
                "recognizedAmount": 25.5, "finalAmount": 25.5, "amountSource": "价税合计",
                "classificationBasis": "通行费", "confidence": "高", "needsReview": "否",
                "expenseCode": "EXP-DEMO", "saveStatus": "成功", "reportCode": "ER-DEMO", "notes": "",
            }],
        }
        with tempfile.TemporaryDirectory() as folder:
            input_path = Path(folder) / "review.json"
            output_path = Path(folder) / "review.xlsx"
            input_path.write_text(json.dumps(review, ensure_ascii=False), encoding="utf-8")
            result = subprocess.run(
                [sys.executable, str(ROOT / "scripts" / "build_review_workbook.py"), str(input_path), str(output_path)],
                check=True, capture_output=True, text=True, encoding="utf-8",
                env={**os.environ, "PYTHONIOENCODING": "utf-8"},
            )
            self.assertIn('"formulaCount"', result.stdout)
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["汇总", "票据明细", "类别核对"])
            self.assertEqual(workbook["票据明细"]["D3"].value, "001234567890")
            self.assertEqual(workbook["票据明细"]["D3"].number_format, "@")
            self.assertTrue(str(workbook["汇总"]["B5"].value).startswith("=SUM("))
            self.assertEqual(workbook["类别核对"]["D3"].value, "=C3-B3")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
