#!/usr/bin/env python3
"""Build a portable Excel review workbook from reimbursement review JSON."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


TITLE_FILL = PatternFill("solid", fgColor="17365D")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=16)
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="17365D")
THIN_BORDER = Border(bottom=Side(style="thin", color="9EADBA"))
CURRENCY_FORMAT = "#,##0.00"
DETAIL_HEADERS = [
    "文件名", "格式", "材料类型", "发票号码", "自动建议分类", "用户确认分类", "归属单据",
    "识别金额", "最终报销金额", "金额来源", "分类依据", "置信度", "是否需核对",
    "汇联易费用编号", "保存状态", "报销单号", "备注",
]


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _style_title(sheet: Any, end_column: int, title: str) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, title)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26
    sheet.sheet_view.showGridLines = False


def _style_header(cells: Any) -> None:
    for cell in cells:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _category_names(data: dict[str, Any]) -> list[str]:
    result: list[str] = []
    for item in data.get("categories") or []:
        value = str(item.get("expenseType") or "").strip()
        if value and value not in result:
            result.append(value)
    for item in data.get("rows") or []:
        value = str(item.get("confirmedCategory") or item.get("suggestedCategory") or "").strip()
        if value and value not in result:
            result.append(value)
    return result


def build_workbook(data: dict[str, Any]) -> Workbook:
    rows = data.get("rows") or []
    categories = data.get("categories") or []
    metadata = data.get("metadata") or {}
    category_names = _category_names(data)
    category_by_name = {str(item.get("expenseType") or ""): item for item in categories}
    detail_end = max(len(rows) + 2, 3)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总"
    details = workbook.create_sheet("票据明细")
    category_sheet = workbook.create_sheet("类别核对")

    _style_title(summary, 6, "报销分类与金额核对")
    summary_rows = [
        ("报销单", "，".join(str(value) for value in metadata.get("reportCodes") or [] if value)),
        ("汇联易回读总额", _number(metadata.get("reportTotal"))),
        ("明细最终金额", f"=SUM('票据明细'!I3:I{detail_end})"),
        ("待确认数量", f'=COUNTIF(\'票据明细\'!M3:M{detail_end},"是")'),
        ("保存失败数量", int(metadata.get("failedExpenses") or 0)),
    ]
    for row_index, values in enumerate(summary_rows, start=3):
        summary.cell(row_index, 1, values[0])
        summary.cell(row_index, 2, values[1])
    _style_header(summary[3][0:1] + summary[4][0:1] + summary[5][0:1] + summary[6][0:1] + summary[7][0:1])
    summary["B4"].number_format = CURRENCY_FORMAT
    summary["B5"].number_format = CURRENCY_FORMAT
    for column, header in enumerate(("类别", "申请金额", "报销金额", "差额", "票据数量"), start=1):
        summary.cell(9, column, header)
    _style_header(summary[9][0:5])
    for offset, name in enumerate(category_names, start=10):
        source = category_by_name.get(name, {})
        summary.cell(offset, 1, name)
        summary.cell(offset, 2, _number(source.get("applicationAmount")))
        summary.cell(offset, 3, f"=SUMIF('票据明细'!F$3:F${detail_end},A{offset},'票据明细'!I$3:I${detail_end})")
        summary.cell(offset, 4, f"=C{offset}-B{offset}")
        summary.cell(offset, 5, f"=COUNTIF('票据明细'!F$3:F${detail_end},A{offset})")
        for column in range(2, 5):
            summary.cell(offset, column).number_format = CURRENCY_FORMAT
    summary.column_dimensions["A"].width = 24
    for column in "BCDEF":
        summary.column_dimensions[column].width = 18

    _style_title(details, len(DETAIL_HEADERS), "票据明细")
    details.append(DETAIL_HEADERS)
    _style_header(details[2])
    for item in rows:
        details.append([
            item.get("fileName") or "",
            item.get("format") or "",
            item.get("documentType") or "",
            str(item.get("invoiceNumber") or ""),
            item.get("suggestedCategory") or "",
            item.get("confirmedCategory") or "",
            item.get("reportGroup") or "",
            item.get("recognizedAmount"),
            item.get("finalAmount"),
            item.get("amountSource") or "",
            item.get("classificationBasis") or "",
            item.get("confidence") or "",
            item.get("needsReview") or "",
            str(item.get("expenseCode") or ""),
            item.get("saveStatus") or "",
            str(item.get("reportCode") or ""),
            item.get("notes") or "",
        ])
    for row_index in range(3, len(rows) + 3):
        for column in (4, 14, 16):
            details.cell(row_index, column).number_format = "@"
        for column in (8, 9):
            details.cell(row_index, column).number_format = CURRENCY_FORMAT
        for cell in details[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if rows:
        table = Table(displayName="InvoiceReviewTable", ref=f"A2:Q{len(rows) + 2}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        details.add_table(table)
        details.conditional_formatting.add(
            f"M3:M{len(rows) + 2}",
            FormulaRule(formula=["M3=\"是\""], fill=PatternFill("solid", fgColor="FFF2CC"), font=Font(color="9C6500", bold=True)),
        )
        details.conditional_formatting.add(
            f"O3:O{len(rows) + 2}",
            FormulaRule(formula=['ISNUMBER(SEARCH("失败",O3))'], fill=PatternFill("solid", fgColor="FCE4D6"), font=Font(color="C00000", bold=True)),
        )
        if category_names and len(",".join(category_names)) <= 250 and not any("," in name for name in category_names):
            validation = DataValidation(type="list", formula1='"' + ",".join(category_names) + '"')
            details.add_data_validation(validation)
            validation.add(f"F3:F{len(rows) + 2}")
    details.freeze_panes = "A3"
    widths = {"A": 34, "B": 10, "C": 12, "D": 24, "E": 18, "F": 18, "G": 14, "H": 14, "I": 16,
              "J": 24, "K": 24, "L": 10, "M": 12, "N": 22, "O": 14, "P": 20, "Q": 28}
    for column, width in widths.items():
        details.column_dimensions[column].width = width

    _style_title(category_sheet, 6, "类别核对")
    category_sheet.append(["类别", "申请金额", "报销金额", "差额", "票据数量", "未匹配数量"])
    _style_header(category_sheet[2])
    for row_index, name in enumerate(category_names, start=3):
        source = category_by_name.get(name, {})
        category_sheet.cell(row_index, 1, name)
        category_sheet.cell(row_index, 2, _number(source.get("applicationAmount")))
        category_sheet.cell(row_index, 3, f"=SUMIF('票据明细'!F$3:F${detail_end},A{row_index},'票据明细'!I$3:I${detail_end})")
        category_sheet.cell(row_index, 4, f"=C{row_index}-B{row_index}")
        category_sheet.cell(row_index, 5, f"=COUNTIF('票据明细'!F$3:F${detail_end},A{row_index})")
        category_sheet.cell(row_index, 6, f'=COUNTIFS(\'票据明细\'!F$3:F${detail_end},A{row_index},\'票据明细\'!O$3:O${detail_end},"未匹配")')
        for column in range(2, 5):
            category_sheet.cell(row_index, column).number_format = CURRENCY_FORMAT
    category_sheet.freeze_panes = "A3"
    category_sheet.column_dimensions["A"].width = 24
    for column in "BCDEF":
        category_sheet.column_dimensions[column].width = 18

    return workbook


def save_verified_workbook(data: dict[str, Any], output_path: Path) -> dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(data)
    workbook.save(output_path)
    reopened = load_workbook(output_path, data_only=False, read_only=False)
    expected = ["汇总", "票据明细", "类别核对"]
    if reopened.sheetnames != expected:
        raise RuntimeError(f"unexpected worksheets: {reopened.sheetnames}")
    formulas = [
        cell.value
        for sheet in reopened.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    invalid = [value for value in formulas if any(error in value for error in ("#REF!", "#VALUE!", "#NAME?"))]
    if invalid:
        raise RuntimeError(f"invalid formulas: {invalid}")
    reopened.close()
    return {"output": str(output_path.resolve()), "worksheets": expected, "formulaCount": len(formulas)}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="review JSON produced by prepare-review")
    parser.add_argument("output", type=Path, help="destination .xlsx path")
    args = parser.parse_args()
    data = json.loads(args.input.read_text(encoding="utf-8"))
    print(json.dumps(save_verified_workbook(data, args.output), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
